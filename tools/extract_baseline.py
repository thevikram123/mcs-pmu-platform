"""
Extract the MCS Phase 3 cost baseline from the three source workbooks into a single
JSON file consumed by the platform.

Sources (read-only, never modified):
  ../gvpr drive data/MCS_Job_Cost_Report_6Year.xlsx   CAPEX + OPEX line-item detail
  ../gvpr drive data/Overhead Cost.xlsx               16-line overhead build-up
  ../gvpr drive data/MCS_Phase3_JCR_Tracker.xlsx      JCR cost codes, vendors, project facts

Output: src/data/baseline.json

Run:  python tools/extract_baseline.py
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT.parent / "gvpr drive data"
OUT = ROOT / "src" / "data" / "baseline.json"

JCR_6YEAR = SRC / "MCS_Job_Cost_Report_6Year.xlsx"
OVERHEAD = SRC / "Overhead Cost.xlsx"
TRACKER = SRC / "MCS_Phase3_JCR_Tracker.xlsx"

# Schedule subtotal row -> (first line-item row, last line-item row).
# Verified against the workbook: every subtotal equals the sum of its own range,
# the ranges do not overlap, and no numeric line item falls outside them.
CAPEX_SCHEDULES = [
    ("A", "Schedule A: Edge Devices", 83, 6, 82),
    ("B", "Schedule B: Primary Data Centre", 142, 86, 141),
    ("C", "Schedule C: Secondary Data Centre", 167, 145, 166),
    ("D", "Schedule D: Central Command Centre (CP Office)", 193, 170, 192),
    ("E", "Schedule E: Police Stations", 213, 196, 212),
    ("F", "Schedule F: Picture Intelligence Unit", 223, 216, 222),
    ("G", "Schedule G: Miscellaneous Costs", 231, 226, 230),
]

OPEX_SCHEDULES = [
    ("H1", "Bandwidth Costs", "Track 1", 26, 6, 25),
    ("H2", "Bandwidth Costs", "Track 2", 40, 29, 39),
    ("I1", "Electricity Cost - Edge Level", "Track 1", 44, 43, 43),
    ("I2", "Electricity Cost - Edge Level", "Track 2", 50, 47, 49),
    ("J1", "O&M for IT / Non-IT Infrastructure", "Track 1", 109, 53, 108),
    ("J2", "O&M for IT / Non-IT Infrastructure", "Track 2", 238, 112, 237),
    ("K", "Managed Hosting Costs (Tier III Data Centre)", "Shared", 246, 241, 245),
    ("L", "Managerial & Technical Manpower Cost", "Shared", 264, 249, 263),
    ("M", "AMC & Operational Cost - Vehicle Tracking System", "Shared", 270, 267, 269),
    ("N", "Cost to Connect Other Establishments", "Shared", 274, 273, 273),
    ("O1", "Operational Manpower Costs", "Track 1", 282, 277, 281),
    ("O2", "Operational Manpower Costs", "Track 2", 290, 285, 289),
]

# CAPEX detail column indices (1-based)
C_NUM, C_DESC, C_QTY, C_UNIT, C_PHASE, C_RR, C_CAT, C_OEM, C_RATE, C_AMT = 2, 3, 4, 5, 6, 7, 8, 9, 10, 11
# OPEX detail column indices
O_NUM, O_DESC, O_QTY, O_UNIT, O_PHASE, O_RR, O_OEM, O_CAT = 2, 3, 4, 5, 6, 7, 8, 9
O_Y1 = 10  # J..O = years 1..6


def num(v):
    """Numeric cell value, or None. Text such as 'NA' becomes None."""
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def text(v):
    if v is None:
        return None
    s = str(v).strip()
    return re.sub(r"\s*\n\s*", " ", s) or None


def slug(*parts):
    raw = "-".join(str(p) for p in parts if p is not None)
    return re.sub(r"[^a-z0-9]+", "-", raw.lower()).strip("-")


def extract_capex(wb):
    ws = wb["CAPEX BOQ - Detail"]
    items, schedules = [], []
    for code, name, total_row, first, last in CAPEX_SCHEDULES:
        count = 0
        for r in range(first, last + 1):
            amt = num(ws.cell(r, C_AMT).value)
            if amt is None or amt == 0:
                continue
            # Subtotal rows carry "TOTAL - Schedule ..." in the # column (B), not in
            # the description; guarding on B keeps genuine line items whose own
            # description begins with "Total" (e.g. "Total MS SOL 2022 Enterprise").
            marker = text(ws.cell(r, C_NUM).value) or ""
            if marker.upper().startswith("TOTAL"):
                continue
            desc = text(ws.cell(r, C_DESC).value) or f"(unnamed line, row {r})"
            qty = num(ws.cell(r, C_QTY).value)
            rate = num(ws.cell(r, C_RATE).value)
            # Some rows carry only a lump amount with no qty/rate pair; model them
            # as qty 1 at the full amount so the qty x rate slider still works.
            if qty is None or rate is None or qty == 0:
                qty, rate = 1.0, amt
            items.append(
                {
                    "id": f"capex-{code}-{r}",
                    "row": r,
                    "schedule": code,
                    "num": text(ws.cell(r, C_NUM).value),
                    "description": desc,
                    "qty": qty,
                    "unit": text(ws.cell(r, C_UNIT).value),
                    "phase": text(ws.cell(r, C_PHASE).value),
                    "rr": text(ws.cell(r, C_RR).value),
                    "category": text(ws.cell(r, C_CAT).value),
                    "oem": text(ws.cell(r, C_OEM).value) or "Unassigned",
                    "unitRate": rate,
                    "amountExGst": amt,
                }
            )
            count += 1
        schedules.append(
            {
                "id": code,
                "name": name,
                "kind": "capex",
                "track": "Track 1" if code == "A" else "Track 2",
                "itemCount": count,
                "sourceTotal": num(ws.cell(total_row, C_AMT).value),
            }
        )
    return items, schedules


def extract_opex(wb):
    ws = wb["OPEX BOQ - Detail"]
    items, schedules = [], []
    for code, name, track, total_row, first, last in OPEX_SCHEDULES:
        count = 0
        for r in range(first, last + 1):
            years = [num(ws.cell(r, O_Y1 + i).value) for i in range(6)]
            if not any(y for y in years if y):
                continue
            marker = text(ws.cell(r, O_NUM).value) or ""
            if marker.upper().startswith("TOTAL"):
                continue
            desc = text(ws.cell(r, O_DESC).value) or f"(unnamed line, row {r})"
            # 'NA' text in years 2-6 (three J1 rows) survives as null and is
            # rendered as NA in the UI rather than being silently zeroed.
            na = [
                i
                for i in range(6)
                if years[i] is None and text(ws.cell(r, O_Y1 + i).value) is not None
            ]
            items.append(
                {
                    "id": f"opex-{code}-{r}",
                    "row": r,
                    "schedule": code,
                    "track": track,
                    "num": text(ws.cell(r, O_NUM).value),
                    "description": desc,
                    "qty": num(ws.cell(r, O_QTY).value),
                    "unit": text(ws.cell(r, O_UNIT).value),
                    "phase": text(ws.cell(r, O_PHASE).value),
                    "rr": text(ws.cell(r, O_RR).value),
                    "category": text(ws.cell(r, O_CAT).value),
                    "oem": text(ws.cell(r, O_OEM).value) or "Unassigned",
                    "years": [y if y is not None else 0.0 for y in years],
                    "naYears": na,
                }
            )
            count += 1
        schedules.append(
            {
                "id": code,
                "name": name,
                "kind": "opex",
                "track": track,
                "itemCount": count,
                "sourceTotal": sum(
                    num(ws.cell(total_row, O_Y1 + i).value) or 0 for i in range(6)
                ),
                "sourceYears": [num(ws.cell(total_row, O_Y1 + i).value) or 0 for i in range(6)],
            }
        )
    return items, schedules


def extract_overhead():
    wb = openpyxl.load_workbook(OVERHEAD, data_only=True)
    ws = wb["Sheet1"]
    items = []
    for r in range(3, 19):
        desc = text(ws.cell(r, 3).value)
        monthly = num(ws.cell(r, 4).value)
        if desc is None or monthly is None:
            continue
        years = [num(ws.cell(r, 5 + i).value) or 0.0 for i in range(6)]
        # Recover the year-on-year escalation the source applied, rather than
        # hardcoding it: rents 10%/yr, salaries 5%/yr, electricity 2% on
        # alternating years, and five flat lines at 0%.
        esc = [
            round(years[i + 1] / years[i] - 1, 6) if years[i] else 0.0 for i in range(5)
        ]
        items.append(
            {
                "id": f"overhead-{r}",
                "row": r,
                "schedule": "P",
                "sr": num(ws.cell(r, 2).value),
                "description": desc,
                "monthlyY1": monthly,
                "escPattern": esc,
                "sourceYears": years,
            }
        )
    total = num(ws.cell(19, 5).value)
    return items, sum(num(ws.cell(19, 5 + i).value) or 0 for i in range(6)), total


def extract_tracker():
    wb = openpyxl.load_workbook(TRACKER, data_only=True)

    cover = wb["Cover"]
    facts = {}
    for r in range(9, 16):
        k, v = text(cover.cell(r, 2).value), text(cover.cell(r, 3).value)
        if k and v:
            facts[k] = v
    facts["Tender Ref"] = text(cover.cell(4, 2).value)
    facts["Project"] = text(cover.cell(2, 2).value)

    summary = wb["JCR Summary"]
    tcv = num(summary.cell(10, 2).value)
    boq_incl = num(summary.cell(11, 2).value)

    cap = wb["JCR Detail - CAPEX"]
    codes = []
    for r in range(4, 13):
        code = text(cap.cell(r, 1).value)
        if not code or code == "TOTAL":
            continue
        codes.append(
            {
                "code": code,
                "description": text(cap.cell(r, 2).value),
                "track": text(cap.cell(r, 3).value),
                "kind": "capex",
                "budgetExGst": num(cap.cell(r, 10).value) or 0.0,
            }
        )

    op = wb["JCR Detail - OPEX"]
    for r in range(5, 18):
        code = text(op.cell(r, 1).value)
        if not code or code == "TOTAL":
            continue
        codes.append(
            {
                "code": code,
                "description": text(op.cell(r, 2).value),
                "track": text(op.cell(r, 3).value),
                "kind": "overhead" if code == "P" else "opex",
                "budgetExGst": num(op.cell(r, 82).value) or 0.0,  # CD = Original Budget
            }
        )

    ven = wb["JCR by Vendor"]
    vendors, section = [], None
    for r in range(1, 61):
        a, b = text(ven.cell(r, 1).value), num(ven.cell(r, 2).value)
        if a and b is None:
            if "CAPEX by OEM" in a:
                section = "capex"
            elif "OPEX by OEM" in a:
                section = "opex"
            continue
        if a and b is not None and section:
            vendors.append({"oem": a, "kind": section, "budgetExGst": b})

    return facts, tcv, boq_incl, codes, vendors


def main():
    for p in (JCR_6YEAR, OVERHEAD, TRACKER):
        if not p.exists():
            sys.exit(f"Source workbook not found: {p}")

    wb = openpyxl.load_workbook(JCR_6YEAR, data_only=True)
    capex_items, capex_schedules = extract_capex(wb)
    opex_items, opex_schedules = extract_opex(wb)
    overhead_items, overhead_total, overhead_y1 = extract_overhead()
    facts, tcv, boq_incl, jcr_codes, vendors = extract_tracker()

    capex_total = sum(i["amountExGst"] for i in capex_items)
    opex_total = sum(sum(i["years"]) for i in opex_items)

    baseline = {
        "meta": {
            "generatedBy": "tools/extract_baseline.py",
            "sources": [JCR_6YEAR.name, OVERHEAD.name, TRACKER.name],
            "horizonYears": 6,
            "currency": "INR",
        },
        "projectFacts": facts,
        "contract": {"tcvInclGst": tcv, "boqBudgetInclGst": boq_incl},
        "schedules": capex_schedules
        + opex_schedules
        + [
            {
                "id": "P",
                "name": "Overhead Cost (Guest House, Office, Warehouse, Vehicles, Core Manpower)",
                "kind": "overhead",
                "track": "Shared",
                "itemCount": len(overhead_items),
                "sourceTotal": overhead_total,
            }
        ],
        "capexItems": capex_items,
        "opexItems": opex_items,
        "overheadItems": overhead_items,
        "jcrCostCodes": jcr_codes,
        "vendors": vendors,
        "checksums": {
            "capexExGst": capex_total,
            "opexExGst": opex_total,
            "overheadExGst": overhead_total,
            "overheadLockTarget": 500_000_000.0,
            "overheadYear1": overhead_y1,
            "projectTotalExGst": capex_total + opex_total + overhead_total,
            "gstRate": 0.18,
        },
        # Defects found while auditing the sources. Surfaced in the UI Data Quality
        # panel; deliberately not corrected in the data itself.
        "dataQuality": [
            {
                "id": "opex-summary-gst-shift",
                "severity": "defect",
                "where": "MCS_Job_Cost_Report_6Year.xlsx > 'OPEX Summary' > columns CF, CG",
                "summary": "GST columns are shifted one column left.",
                "detail": (
                    "CF is headed 'GST @ 18%' but links to 'OPEX BOQ - Detail'!P (the Excl-GST "
                    "total), and CG is headed 'Total - Incl. GST' but links to column Q (the GST "
                    "amount). CF19 therefore reads Rs.6,917,661,679.69 where Rs.1,245,179,102.34 "
                    "is meant. Executive Summary is unaffected because it computes GST itself as "
                    "F7*18%. This platform computes GST from the ex-GST base, so its figures and "
                    "its Excel export are correct."
                ),
            },
            {
                "id": "opex-o1-row279-year3",
                "severity": "anomaly",
                "where": "MCS_Job_Cost_Report_6Year.xlsx > 'OPEX BOQ - Detail'!L279",
                "summary": "Schedule O1 'Manpower in Mobile Vans' breaks its 4% escalation in Year 3.",
                "detail": (
                    "Year 3 reads 3,959,612 against a Year 2 of 5,643,938, then jumps back to "
                    "6,104,484 in Year 4. Every other row in O1/O2 escalates cleanly, and this "
                    "row's Track 2 twin (row 286) reads 5,869,696 in Year 3, which is exactly "
                    "5,643,938 x 1.04. The figure appears to understate by about Rs.19.1 lakh. "
                    "Carried through as-is; a decision for the BOQ owner."
                ),
            },
        ],
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(baseline, indent=1, ensure_ascii=False), encoding="utf-8")

    print(f"wrote {OUT.relative_to(ROOT)}  ({OUT.stat().st_size / 1024:.0f} KB)")
    print(f"  CAPEX     {len(capex_items):>4} items  Rs.{capex_total:>18,.2f}")
    print(f"  OPEX      {len(opex_items):>4} items  Rs.{opex_total:>18,.2f}")
    print(f"  Overhead  {len(overhead_items):>4} items  Rs.{overhead_total:>18,.2f}")
    print(f"  TOTAL                 Rs.{capex_total + opex_total + overhead_total:>18,.2f}")
    print(f"  JCR codes {len(jcr_codes):>4}   vendors {len(vendors)}")

    # Fail loudly if the extraction drifts from the audited workbook totals.
    expect = {
        "capex": 3_294_610_663.98,
        "opex": 6_917_661_679.69,
        "overhead": 532_324_821.19,
    }
    for label, got, want in (
        ("CAPEX", capex_total, expect["capex"]),
        ("OPEX", opex_total, expect["opex"]),
        ("Overhead", overhead_total, expect["overhead"]),
    ):
        if abs(got - want) > 1.0:
            sys.exit(f"FAIL {label}: extracted {got:,.2f}, workbook says {want:,.2f}")
    print("  checksums OK")


if __name__ == "__main__":
    main()
