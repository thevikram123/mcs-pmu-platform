"""
Extract the MCS Phase III milestone schedule into src/data/milestones.json.

Three sources, three different numbering schemes, deliberately kept distinct:

  MSA_Cost_Extract.xlsx > '4. Payment Milestones'
      M1..M29 — the CONTRACTUAL payment milestones: % of TCV, amount payable,
      timeline relative to the T anchors. This is what the client gets invoiced
      against, and its percentages sum to exactly 100%.

  Milestone_Mapped_PO_Tracker_updated.xlsx
      'Original MS' M1..M32 — the OPERATIONAL milestone list with real calendar
      start/end dates, plus 321 purchase-order line items hung off them.
      'New' M0..M14 — a renumbered subset covering only milestones that carry
      PO items.

The MSA is the supreme authority. It post-dates the RFP and the four
corrigendums, and where they differ the MSA governs — the corrigendum milestone
tables in particular use different numbering and percentages and are not used
here.

The MSA defines T4 as "the date of handover of the new control room to the SI"
and fixes no calendar date for it, because it depends on a third party. It does
fix everything downstream: T5 = T4 + 2 months, T6 = T5 + 1 month. So T4 is
carried as a single settable project assumption and M28/M29 derive from it by
the MSA formula. Primavera Time Line (1).xlsx supplies only the default value
for that assumption (it schedules the work from 12 Feb 2027, just after Track 2
go-live); nothing else in the platform depends on it.

Run:  python tools/extract_milestones.py
"""

from __future__ import annotations

import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT.parent / "gvpr drive data"
OUT = ROOT / "src" / "data" / "milestones.json"

MSA_COST = ROOT.parent / "MSA_Cost_Extract.xlsx"
PO_TRACKER = SRC / "Milestone_Mapped_PO_Tracker_updated.xlsx"


def text(v):
    if v is None:
        return None
    s = re.sub(r"\s*\n\s*", " ", str(v).strip())
    return s or None


def num(v):
    return v if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def iso(v):
    """Excel date -> ISO date string, or None."""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


# --------------------------------------------------------------- payments

def extract_payments():
    wb = openpyxl.load_workbook(MSA_COST, data_only=True)
    ws = wb["4. Payment Milestones"]
    rows = []
    for r in range(5, 34):
        mid = text(ws.cell(r, 1).value)
        if not mid or not mid.startswith("M"):
            continue
        pct = num(ws.cell(r, 5).value)
        times = num(ws.cell(r, 6).value)
        amount = num(ws.cell(r, 7).value)
        rows.append(
            {
                "id": mid,
                "track": text(ws.cell(r, 2).value) or "Common",
                "name": text(ws.cell(r, 3).value),
                "timeline": text(ws.cell(r, 4).value),
                "pctOfTcv": pct,          # per event; None where no separate %
                "times": int(times) if times else None,
                "amountInclGst": amount,  # total across all occurrences
                "deliverable": text(ws.cell(r, 9).value),
            }
        )
    check = num(ws.cell(34, 7).value)  # printed arithmetic check row
    return rows, check


# ------------------------------------------------------- operational plan

MS_HEADER = re.compile(r"^(?:(M\d+)\s+)?\[(M\d+)\]\s*(.*)$")


def extract_schedule_and_pos():
    wb = openpyxl.load_workbook(PO_TRACKER, data_only=True)
    ws = wb["Milestone-Mapped PO Tracker"]

    schedule, po_items = [], []
    for r in range(3, ws.max_row + 1):
        a = text(ws.cell(r, 1).value)
        b = text(ws.cell(r, 2).value)
        if not a:
            continue

        m = MS_HEADER.match(a)
        if m and not b:
            new_id, orig_id, name = m.groups()
            start, end = iso(ws.cell(r, 8).value), iso(ws.cell(r, 9).value)
            schedule.append(
                {
                    "id": orig_id,
                    "newId": new_id,
                    "name": name.strip(),
                    "timelineText": text(ws.cell(r, 7).value),
                    "start": start,
                    "end": end,
                    "durationDays": num(ws.cell(r, 10).value),
                }
            )
        elif b:
            po_items.append(
                {
                    "milestoneId": b,
                    "newId": a,
                    "milestone": text(ws.cell(r, 3).value),
                    "category": text(ws.cell(r, 4).value) or "Unassigned",
                    "itemNo": text(ws.cell(r, 5).value),
                    "description": text(ws.cell(r, 6).value),
                    "poTargetDate": iso(ws.cell(r, 11).value)
                    or text(ws.cell(r, 11).value),
                    "status": text(ws.cell(r, 12).value) or "Pending",
                    "remarks": text(ws.cell(r, 13).value),
                }
            )
    return schedule, po_items


def main():
    for p in (MSA_COST, PO_TRACKER):
        if not p.exists():
            sys.exit(f"Source workbook not found: {p}")

    payments, printed_check = extract_payments()
    schedule, po_items = extract_schedule_and_pos()

    total_amount = sum(p["amountInclGst"] or 0 for p in payments)
    total_pct = sum((p["pctOfTcv"] or 0) * (p["times"] or 0) for p in payments)

    # Calendar anchors, read off the operational schedule.
    by_id = {s["id"]: s for s in schedule}
    anchors = {
        "T": by_id.get("M4", {}).get("end"),      # contract signing
        "T1": by_id.get("M6", {}).get("end"),     # handover from existing SI
        "T2": by_id.get("M28", {}).get("end"),    # Track 2 project go-live
    }

    # Read directly off the 38 MSA page scans in gvpr drive data/msa_pages and
    # cross-checked against the '4. Payment Milestones' extract: every milestone
    # percentage above was confirmed against the signed pages.
    governance = {
        "anchorDefinitions": [
            {"anchor": "T", "definition": "Contract signing / date of award", "page": "8"},
            {"anchor": "T1", "definition": "T + 3 months — handover from the existing SI", "page": "8"},
            {"anchor": "T2", "definition": "T + 9 months — Project Go-Live of Track 2", "page": "14"},
            {"anchor": "T3", "definition": "T2 + 3 months", "page": "14"},
            {
                "anchor": "T4",
                "definition": "Date the new control room is handed over to the SI — not fixed to a calendar date in the MSA",
                "page": "15",
            },
            {
                "anchor": "T5",
                "definition": "T4 + 2 months (RFP Vol.1 said T4 + 3 months; amended by Corrigendum 3 and carried into the MSA)",
                "page": "15",
            },
            {"anchor": "T6", "definition": "T5 + 1 month", "page": "15"},
        ],
        "clauses": [
            {
                "title": "Payment release condition",
                "ref": "p.8",
                "text": "Payment is released only on satisfactory acceptance of the deliverables for each task or activity.",
            },
            {
                "title": "SLA penalty",
                "ref": "Note A, p.16",
                "text": "One SLA mark equals 1% of the penalty amount, adjusted proportionately across the quarterly O&M payments under both Track 1 and Track 2. No separate LD cap is stated in the milestone schedule.",
            },
            {
                "title": "Overall liability cap",
                "ref": "Cl. 20, p.24",
                "text": "MSI liability shall not exceed 50% of the total contract value (Rs.1,049.46 Cr). Risk-purchase liability is separately capped at 20% of TCV (Cl. 19, p.23).",
            },
            {
                "title": "Deemed acceptance",
                "ref": "Cl. N (Survivability), p.36",
                "text": "A deliverable is deemed accepted if the department does not respond within 30 days of delivery, provided it went to the designated officer and was flagged in at least three weekly progress reports.",
            },
            {
                "title": "Performance Bank Guarantee",
                "ref": "Cl. 5.2(a) & Cl. 8, p.5 & 17",
                "text": "3% of TCV inclusive of GST — Rs.62.97 Cr, submitted 04 May 2026.",
            },
        ],
        "resolutionOfT4": (
            "Checked across the whole document chain. RFP Vol.2 requires 'shifting of the "
            "Command and Control Setup to the New Integrated Control Room and Operations Centre "
            "at the CP Office (once ready)'. Corrigendums 2, 3 and 4 and the signed MSA all "
            "define T4 identically as 'the date of handover of the new control room to the SI'. "
            "No document anywhere fixes a calendar date for it: the anchor is condition-based, "
            "triggered by the department handing over the room, and that is the resolution rather "
            "than an omission. Everything downstream is fixed, so setting T4 dates M28 and M29 "
            "exactly."
        ),
        "amendments": [
            {
                "item": "T5 definition",
                "rfp": "T4 + 3 months (RFP Vol.1)",
                "final": "T4 + 2 months",
                "via": "Corrigendum 3, carried into the signed MSA",
            },
            {
                "item": "Milestone numbering and percentages",
                "rfp": "Corrigendum 2 lists M9-M17 on a different numbering with different percentages",
                "final": "M1-M29 as printed in the MSA",
                "via": "MSA supersedes the corrigendum tables",
            },
        ],
        "verification": (
            "The 29 payment milestones and their percentages were read independently off the "
            "38 signed MSA page scans and matched the spreadsheet extract on every checked "
            "value, including M2 at 2.5%, M4 at 2.25% x 24 quarters, M15 and M19 at 4.0%, "
            "M26 at 0.7% x 20 quarters, and a total of 100% of TCV."
        ),
    }

    data = {
        "meta": {
            "generatedBy": "tools/extract_milestones.py",
            "sources": [MSA_COST.name, PO_TRACKER.name],
        },
        "anchors": anchors,
        # T4 is not fixed by any document. Default taken from the Primavera plan,
        # which starts the New CCC work on 12 Feb 2027 — six days after Track 2
        # go-live, i.e. the earliest point the new control room plausibly exists.
        # Settable in the UI; M28 and M29 follow by the MSA formula.
        "assumedT4": "2027-02-12",
        "t4Basis": (
            "Condition-based in every source document — no calendar date is fixed anywhere. "
            "Defaulted to the Primavera plan's start for the New CCC work, six days after "
            "Track 2 go-live. Set it to the actual handover date once known."
        ),
        "governance": governance,
        "payments": payments,
        "schedule": schedule,
        "poItems": po_items,
        "checksums": {
            "paymentTotalInclGst": total_amount,
            "paymentPctTotal": total_pct,
            "printedCheck": printed_check,
            "paymentCount": len(payments),
            "scheduleCount": len(schedule),
            "poItemCount": len(po_items),
        },
        "dataQuality": [
            {
                "id": "milestone-t4-assumption",
                "severity": "anomaly",
                "where": "MSA Cl. 7 payment schedule — anchor T4",
                "summary": "T4 is a project assumption, not a contractual date.",
                "detail": (
                    "The MSA defines T4 as the date the new control room is handed over to the "
                    "SI and fixes no calendar date, because the handover is the department's to "
                    "make. This was checked against the full chain: RFP Vol.2 makes the shift "
                    "conditional on the room being 'once ready', and Corrigendums 2, 3 and 4 "
                    "repeat the MSA wording verbatim. The anchor is condition-based by design. "
                    "Everything downstream is fixed: T5 = T4 + 2 months and T6 = T5 + 1 month, "
                    "so M28 (Rs.12.59 Cr) and M29 (Rs.8.40 Cr) follow automatically once T4 is "
                    "set. The platform carries T4 as a single settable assumption, defaulted to "
                    "12 Feb 2027 from the Primavera plan — six days after Track 2 go-live, the "
                    "earliest point the new control room plausibly exists. Change T4 on the "
                    "Contract terms tab and both milestones move with it."
                ),
            },
            {
                "id": "milestone-numbering-differs",
                "severity": "anomaly",
                "where": "MSA payment milestones vs Milestone-Mapped PO Tracker",
                "summary": "Three different milestone numbering schemes are in use.",
                "detail": (
                    "The MSA pays against M1-M29. The PO tracker's 'Original MS' column runs "
                    "M1-M32 and starts earlier (LoA issuance, PBG submission), so the same "
                    "label means different things in the two documents — MSA M5 is a Track 1 "
                    "supply milestone, tracker M5 is submission of the Advance Bank Guarantee. "
                    "The tracker's 'New' column renumbers again as M0-M14 for the subset that "
                    "carries purchase orders. This platform keeps all three and never merges "
                    "them on the label alone."
                ),
            },
        ],
    }

    OUT.write_text(json.dumps(data, indent=1, ensure_ascii=False), encoding="utf-8")
    print(f"wrote {OUT.relative_to(ROOT)}  ({OUT.stat().st_size / 1024:.0f} KB)")
    print(f"  payment milestones {len(payments):>4}   Rs.{total_amount:>18,.2f}")
    print(f"  pct of TCV total        {total_pct * 100:>8.2f}%")
    print(f"  operational milestones {len(schedule):>3}")
    print(f"  PO line items      {len(po_items):>4}")
    print(f"  anchors            {anchors}")

    if abs(total_pct - 1.0) > 0.0005:
        sys.exit(f"FAIL: milestone percentages total {total_pct:.4%}, expected 100%")
    if printed_check and abs(total_amount - printed_check) > 1:
        sys.exit(f"FAIL: amounts total {total_amount:,.2f} vs printed {printed_check:,.2f}")
    print("  checksums OK")


if __name__ == "__main__":
    main()
